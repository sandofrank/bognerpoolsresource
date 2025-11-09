#!/usr/bin/env python3
import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from copy import copy

print('Reorganizing price list with full formatting preservation...')

# Read the JSON data
with open('public/price-data.json', 'r') as f:
    current_data = json.load(f)

# Load the original workbook with full formatting
original_wb = openpyxl.load_workbook('Archive/Bogner Price List 11.11.25.xlsx')
original_ws = original_wb.active

# Load again with data_only to get computed values for formulas
original_wb_values = openpyxl.load_workbook('Archive/Bogner Price List 11.11.25.xlsx', data_only=True)
original_ws_values = original_wb_values.active

# Create new workbook
new_wb = openpyxl.Workbook()
new_ws = new_wb.active
new_ws.title = 'Price List Reorganized'

# Helper function to find row by text in column C
def find_row_by_text(search_text, search_type='exact'):
    for row_idx, row in enumerate(original_ws.iter_rows(min_col=3, max_col=3), start=1):
        cell = row[0]
        if cell.value:
            cell_value = str(cell.value).strip()
            search_value = str(search_text).strip()

            if search_type == 'exact' and cell_value == search_value:
                return row_idx
            elif search_type == 'contains' and search_value in cell_value:
                return row_idx
    return -1

# Helper function to copy a row with all formatting
def copy_row(source_row_idx, target_row_idx):
    source_row = original_ws[source_row_idx]

    for col_idx, source_cell in enumerate(source_row, start=1):
        if col_idx > 3:  # Only copy columns A, B, C
            break

        target_cell = new_ws.cell(row=target_row_idx, column=col_idx)
        source_cell_value = original_ws_values.cell(row=source_row_idx, column=col_idx)

        # Copy value (use value workbook to avoid formulas)
        if source_cell_value.value is not None:
            target_cell.value = source_cell_value.value
        elif source_cell.value is not None and not (isinstance(source_cell.value, str) and source_cell.value.startswith('=')):
            target_cell.value = source_cell.value

        # Copy formatting carefully
        if source_cell.has_style:
            try:
                if source_cell.font:
                    target_cell.font = copy(source_cell.font)
            except:
                pass
            try:
                if source_cell.border:
                    target_cell.border = copy(source_cell.border)
            except:
                pass
            try:
                if source_cell.fill:
                    target_cell.fill = copy(source_cell.fill)
            except:
                pass
            try:
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
            except:
                pass
            try:
                if source_cell.alignment:
                    target_cell.alignment = copy(source_cell.alignment)
            except:
                pass

# Helper function to get all rows for a section
def get_section_rows(section_name):
    start_row = find_row_by_text(section_name, 'exact')
    if start_row == -1:
        return []

    rows = [start_row]

    # Continue until we hit another section header
    for row_idx in range(start_row + 1, original_ws.max_row + 1):
        cell_c = original_ws.cell(row=row_idx, column=3)
        cell_a = original_ws.cell(row=row_idx, column=1)

        # Stop if we hit a new section (no cost/unit and not a note)
        if cell_c.value and not cell_a.value:
            text = str(cell_c.value).strip()
            if not text.startswith('Note'):
                break

        rows.append(row_idx)

        # Stop if we've gone far enough and hit empty cells
        if not cell_c.value and not cell_a.value:
            next_cell = original_ws.cell(row=row_idx + 1, column=3)
            if not next_cell.value:
                break

    return rows

# Build reorganized data structure
reorganized_data = {
    'categories': []
}

# NEW CONSTRUCTION
new_construction = {'name': 'NEW CONSTRUCTION', 'sections': []}

# Phase headers and sections as defined in the JS script
sections_order = [
    {'type': 'phase', 'name': 'PHASE 1: PROJECT PLANNING & SITE PREPARATION'},
    {'type': 'section', 'name': 'Pool Base Cost'},
    {'type': 'section', 'name': 'Pool Depths'},
    {'type': 'section', 'name': 'Zone Charges'},
    {'type': 'section', 'name': 'Structural Engineering'},
    {'type': 'section', 'name': 'Excavation'},

    {'type': 'phase', 'name': 'PHASE 2: POOL SHELL CONSTRUCTION'},
    {'type': 'section', 'name': 'Shotcrete'},
    {'type': 'section', 'name': 'Raised Bond Beam'},
    {'type': 'section', 'name': 'Wing Walls'},
    {'type': 'section', 'name': 'Vanishing Edge'},
    {'type': 'section', 'name': 'Spa Base Cost'},
    {'type': 'section', 'name': 'Spa Extras'},
    {'type': 'section', 'name': 'Spa Walls'},
    {'type': 'section', 'name': 'Spa Only'},

    {'type': 'phase', 'name': 'PHASE 3: PLUMBING SYSTEMS'},
    {'type': 'section', 'name': 'Pool Plumbing'},
    {'type': 'section', 'name': 'Spa Plumbing'},
    {'type': 'section', 'name': 'In Floor Cleaner'},
    {'type': 'section', 'name': 'Pool Sweeps'},
    {'type': 'section', 'name': 'Drains'},
    {'type': 'section', 'name': 'Gas Lines'},

    {'type': 'phase', 'name': 'PHASE 4: EQUIPMENT INSTALLATION'},
    {'type': 'section', 'name': 'Heaters'},
    {'type': 'section', 'name': 'Sanitizers'},

    {'type': 'phase', 'name': 'PHASE 5: ELECTRICAL SYSTEMS'},
    {'type': 'section', 'name': 'Electrical'},
    {'type': 'section', 'name': 'Remote Controls'},

    {'type': 'phase', 'name': 'PHASE 6: INTERIOR FINISHES'},
    {'type': 'section', 'name': 'Tile'},
    {'type': 'section', 'name': 'Coping'},
    {'type': 'section', 'name': 'Pebble/Quartz/Colored Plaster'},

    {'type': 'phase', 'name': 'PHASE 7: POOL FEATURES & ACCESSORIES'},
    {'type': 'section', 'name': 'Pool Extras'},
    {'type': 'section', 'name': 'Water Features'},
    {'type': 'section', 'name': 'Motorized Pool Cover'},

    {'type': 'phase', 'name': 'PHASE 8: DECKING & HARDSCAPE'},
    {'type': 'section', 'name': 'Decking'},
    {'type': 'section', 'name': 'Footings'},
    {'type': 'section', 'name': 'Step Risers'},
    {'type': 'section', 'name': 'Pavers'},

    {'type': 'phase', 'name': 'PHASE 9: WALLS & STRUCTURES'},
    {'type': 'section', 'name': 'Walls'},
    {'type': 'section', 'name': 'Wall Caps'},
    {'type': 'section', 'name': 'Columns'},
    {'type': 'section', 'name': 'Column Caps'},
    {'type': 'section', 'name': 'Facing and Veneer Not Raised Bond Beam'},

    {'type': 'phase', 'name': 'PHASE 10: OUTDOOR LIVING FEATURES'},
    {'type': 'section', 'name': 'Bbqs'},
    {'type': 'section', 'name': 'Fire Pits and Bowls'},
    {'type': 'section', 'name': 'Real Rock'},
    {'type': 'section', 'name': 'Artificial Rock'},
    {'type': 'section', 'name': 'Solar'},
    {'type': 'section', 'name': 'Alumawood Patio Covers'},

    {'type': 'phase', 'name': 'PHASE 11: FENCING & SAFETY'},
    {'type': 'section', 'name': 'Fencing'},
]

# Copy column widths
for col_idx in range(1, 4):
    col_letter = get_column_letter(col_idx)
    if original_ws.column_dimensions[col_letter].width:
        new_ws.column_dimensions[col_letter].width = original_ws.column_dimensions[col_letter].width

# Make sure column C is wide enough for phase headers
if new_ws.column_dimensions['C'].width and new_ws.column_dimensions['C'].width < 60:
    new_ws.column_dimensions['C'].width = 60

# Start copying rows
current_target_row = 1

# Copy header row (Cost, Factor, Item)
header_row = 1
for row in original_ws.iter_rows(min_row=1, max_row=10):
    if row[0].value == 'Cost' or row[2].value == 'Item':
        header_row = row[0].row
        break

copy_row(header_row, current_target_row)
current_target_row += 1

# Copy "Last Updated" row
last_updated_row = find_row_by_text('Last Updated', 'contains')
if last_updated_row > 0:
    copy_row(last_updated_row, current_target_row)
    current_target_row += 1

# Copy NEW CONSTRUCTION category row
category_row = find_row_by_text('New Construction', 'exact')
print(f'Found New Construction at row: {category_row}')
if category_row > 0:
    copy_row(category_row, current_target_row)
    # Change text to uppercase
    new_ws.cell(row=current_target_row, column=3).value = 'NEW CONSTRUCTION'
    current_target_row += 1
else:
    print('ERROR: Could not find New Construction category!')

# Process all NEW CONSTRUCTION sections
for item in sections_order:
    if item['type'] == 'phase':
        # Copy category row format for phase header
        phase_template_row = find_row_by_text('New Construction', 'exact')
        if phase_template_row <= 0:
            phase_template_row = find_row_by_text('Remodel', 'exact')

        if phase_template_row > 0:
            copy_row(phase_template_row, current_target_row)
            # Clear columns A and B, update column C with phase name
            phase_cell_a = new_ws.cell(row=current_target_row, column=1)
            phase_cell_b = new_ws.cell(row=current_target_row, column=2)
            phase_cell_c = new_ws.cell(row=current_target_row, column=3)

            phase_cell_a.value = None
            phase_cell_b.value = None
            phase_cell_c.value = item['name']

            # Customize phase header formatting using website color scheme (slate-700)
            phase_cell_c.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
            phase_cell_c.fill = PatternFill(start_color='334155', end_color='334155', fill_type='solid')  # slate-700
            phase_cell_c.alignment = Alignment(horizontal='center', vertical='center')

            print(f'Created phase header: {item["name"]}')
            current_target_row += 1
    else:
        # Copy entire section with all its rows
        section_rows = get_section_rows(item['name'])
        if section_rows:
            print(f'Copying section "{item["name"]}" - {len(section_rows)} rows')
            for source_row in section_rows:
                copy_row(source_row, current_target_row)
                current_target_row += 1
        else:
            print(f'WARNING: Section "{item["name"]}" not found!')

# Copy REMODEL category row
remodel_row = find_row_by_text('Remodel', 'exact')
print(f'Found Remodel at row: {remodel_row}')
if remodel_row > 0:
    copy_row(remodel_row, current_target_row)
    # Change text to uppercase
    new_ws.cell(row=current_target_row, column=3).value = 'REMODEL'
    current_target_row += 1

# REMODEL sections (simplified - copy all remodel sections as they are)
remodel_sections = [
    {'type': 'phase', 'name': 'REMODEL PHASE 1: DEMOLITION & PREPARATION'},
    {'type': 'section', 'name': 'Strip Plaster'},
    {'type': 'section', 'name': 'Dump Fees'},
    {'type': 'phase', 'name': 'REMODEL PHASE 2: SURFACE PREPARATION'},
    {'type': 'section', 'name': 'Cut Tile if Tile Stays on Pool and Spa'},
    {'type': 'section', 'name': 'Tile up to Group 5'},
    {'type': 'section', 'name': 'Coping'},
    {'type': 'phase', 'name': 'REMODEL PHASE 3: NEW FINISHES'},
    {'type': 'section', 'name': 'White Plaster'},
    {'type': 'section', 'name': 'Pebble Finish'},
    {'type': 'section', 'name': 'Sparkle Quartz'},
    {'type': 'phase', 'name': 'REMODEL PHASE 4: SYSTEMS UPGRADES'},
    {'type': 'section', 'name': 'Plumbing'},
    {'type': 'section', 'name': 'Electrical'},
    {'type': 'section', 'name': 'Equipment'},
    {'type': 'phase', 'name': 'REMODEL PHASE 5: DECKING'},
    {'type': 'section', 'name': 'Decking'},
]

for item in remodel_sections:
    if item['type'] == 'phase':
        # Copy category row format for phase header
        phase_row = find_row_by_text('Remodel', 'exact')
        if phase_row > 0:
            copy_row(phase_row, current_target_row)
            # Clear columns A and B, update column C with phase name
            phase_cell_a = new_ws.cell(row=current_target_row, column=1)
            phase_cell_b = new_ws.cell(row=current_target_row, column=2)
            phase_cell_c = new_ws.cell(row=current_target_row, column=3)

            phase_cell_a.value = None
            phase_cell_b.value = None
            phase_cell_c.value = item['name']

            # Customize remodel phase header formatting using website color scheme (slate-600 - lighter grey)
            phase_cell_c.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
            phase_cell_c.fill = PatternFill(start_color='475569', end_color='475569', fill_type='solid')  # slate-600
            phase_cell_c.alignment = Alignment(horizontal='center', vertical='center')

            print(f'Created remodel phase header: {item["name"]}')
            current_target_row += 1
    else:
        # Copy entire section with all its rows
        section_rows = get_section_rows(item['name'])
        if section_rows:
            print(f'Copying remodel section "{item["name"]}" - {len(section_rows)} rows')
            for source_row in section_rows:
                copy_row(source_row, current_target_row)
                current_target_row += 1
        else:
            print(f'WARNING: Remodel section "{item["name"]}" not found!')

# Save the new workbook
new_wb.save('Archive/Bogner Price List - Reorganized.xlsx')

print('✓ Reorganization complete with full formatting preserved!')
print(f'✓ File saved to: Archive/Bogner Price List - Reorganized.xlsx')
print(f'✓ Total rows copied: {current_target_row - 1}')
